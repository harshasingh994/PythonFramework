import openpyxl
book = openpyxl.load_workbook(r"C:\Users\SESA548418\Documents\PythonTestData.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)   # get from sheet
print(cell.value)
# sheet.cell(row=2, column=2).value = "test"   # write in sheet
# print(sheet.cell(row=2, column=2).value)

print(sheet.max_column)         # getting max column
print(sheet.max_row)             # getting max row

for i in range(2, sheet.max_row+1):          # for rows
    for j in range(1, sheet.max_column+1):  # for col
        print(sheet.cell(row=i, column=j).value)
        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print(Dict)
